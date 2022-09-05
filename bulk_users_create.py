# -*- coding: utf-8 -*-
#
# created by artemk
# email : kochetkov1985@mail.ru
#

import json
from datetime import datetime
import os.path
import csv
from traceback import print_tb
from jira import JIRA, JIRAError
from logging.handlers import RotatingFileHandler
import logging
import sys
import requests
import config
from openpyxl import load_workbook
import xlsxwriter
from utils import generatepassword, createlogin
import re


rotate = logging.handlers.RotatingFileHandler('bulk_create_users.log', maxBytes=10000000, backupCount=5,
                                              encoding='utf-8')
consoleHandler = logging.StreamHandler(sys.stdout)

logging.basicConfig(format="[%(asctime)s] [%(levelname)8s] --- %(message)s (%(filename)s:%(lineno)s)",
                    level=logging.INFO, handlers=[rotate, consoleHandler])


def jiraConnect(url):
    """
    Connect to JIRA

    :return: JIRA instance
    """
    try:
        jira_options = {'server': url, 'verify': 'certs.pem'}
        jira = JIRA(options=jira_options, basic_auth=(
            config.USERS_ADD_LOGIN, config.USERS_ADD_PASS))
        logging.info(f"Подключение к {config.USERS_ADD_BASE_URL} установлено.")
        return jira
    except JIRAError as error:
        logging.error(error)


def readexcelfile(file):
    """
    Read Excel file

    :return: list of dictionary {FIO: email}
    """
    users_list = []
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[config.EXCEL_WB_NAME]

    for row in range(config.EXCEL_START_ROW, config.EXCEL_END_ROW):
        issue_data = {}

        if ws.cell(row, 1).value == None and ws.cell(row, 2).value == None and ws.cell(row, 3).value == None:
            logging.info(f'End data in Excel in {row} row.')
            break

        issue_data['FIO'] = ws.cell(row, 2).value.strip()
        issue_data['email'] = ws.cell(row, 3).value.strip()

        if issue_data['FIO']:
            users_list.append(issue_data)

    wb.close()
    return users_list


def searchuser(jira, user):
    """
    Search user in JIRA

    :return: True or False
    """
    try:
        users = jira.search_users(user['email'])
        if users:
            return False
        else:
            return True
    except JIRAError as e:
        logging.error(e.text)


def addnewusertojira(jira, user, file):
    """
    Add user in JIRA and write data to file
    login and password for new user generate automatic

    jira: JIRA instance
    user: dict {
        'FIO': FIO,
        'email': email
        }
    file: file to write data
    :return: True or False
    """
    try:
        if searchuser(jira, user):
            login = createlogin(user['FIO'])
            upass = generatepassword()
            try:
                jira.add_user(
                    username=login, email=user['email'], password=upass, fullname=user['FIO'], directoryId=1, application_keys=[])
                jira.add_user_to_group(username=login, group='accept_ext')
                file.write(
                    f" '{user['FIO'].strip()}' [{user['email'].strip()}] : {login} - {upass}\n")
                return True
            except JIRAError as e:
                logging.error(e)
        else:
            logging.error(f"Пользователь {user['FIO']} уже существует !")
            return False
    except Exception as e:
        logging.error(e)


def addusertojira(jira, user):
    """
    Add user in JIRA

    jira: JIRA instance
    user: dict {
        'FIO': FIO,
        'login': login,
        'email': email,
        'upass': upass
        }
    
    :return: True or False
    """

    try:
        if searchuser(jira, user):
            jira.add_user(
                    username=user['login'], email=user['email'], password=user['upass'], fullname=user['FIO'], directoryId=1, application_keys=[])
            jira.add_user_to_group(username=user['login'], group='accept_ext')
            return True
        else:
            return False
    except JIRAError as e:
        logging.error(e)
        return False


def getusersfromcsv(file):
    users_data = []
    with open(file, newline='') as File:
        reader = csv.reader(File)
        for row in reader:
            users_data.append(row)
    return users_data


def getjirausers(jira, file, mode):
    users_data = getusersfromcsv(file)
    count = 0

    if (mode == 'Activate'):
        list_users = set()
        for user in users_data:
            count += 1
            otr_email = re.search(r'\w+@otr.ru', user[2])
            if (otr_email):
                # response["message"] = ""
                list_users.add(user[1])
    
    if (mode=='Deactivate'):
        list_users = list()
        for user in users_data:
            list_users.add(user[2])

    response = {
        "message": ""
    }
    try:
        if (mode=='Activate'):
            print(jira._session.cookies.get_dict())
            response = updatejirauser(jira._session.cookies.get_dict(), list(list_users), mode='Activate')
        elif (mode=='Deactivate'):
            response = updatejirauser(jira._session.cookies.get_dict(), list(list_users))

        if response["message"] == "all users activated":
            logging.info(
                f"Пользователь № {count} [{user[3]}] - активирован !")
        elif response["message"] == "all users deactivated":
            logging.info(
                f"Пользователь № {count} [{user[3]}] - деактивирован !")
        else:
            # logging.error(f"Не удалось обновить пользователя № {count} : [{user[3]}]. Причина : {response['message']}.")
            pass
    except IndexError as err:
        logging.error(f"Data not Found : {err} - {user}")
    except JIRAError as error:
        logging.error(error)


def addjirausersingroup(jira, file):
    users_data = getusersfromcsv(file)
    count = 0

    list_users = set()
    for user in users_data:
        count += 1
        otr_email = re.search(r'\w+@otr.ru', user[2])
        if (otr_email):
            list_users.add(user[1])

    for user in list_users:
        try:
            jira.add_user_to_group(user, 'accept_int')
            logging.info(f'Пользователь с {user} добавлен в группу accept_int.')
        except JIRAError as err:
            logging.error(err)


def updatejirauser(cookie, user, mode='Deactivate'):

    payload = json.dumps({
        "users": user
    })

    # API JIRA Extender plugin
    if mode == 'Deactivate':
        url = f"{config.USERS_ADD_BASE_URL}/rest/extender/1.0/user/deactivate"
    
    if mode == 'Activate':
        url = f"{config.USERS_ADD_BASE_URL}/rest/extender/1.0/user/activate"
    
    response = requests.request(
        "POST",
        url,
        data=payload,
        cookies=cookie
    )

    # logging.info(response.text)
    return response.json()


def writexls(file):
    xsl_file = 'user_data.xlsx'
    doc = xlsxwriter.Workbook(filename=xsl_file)
    doc_ws = doc.add_worksheet('Users_data')
    count = 0

    users_data = parsepassfile(file)

    doc_ws.write(0, 1, "ФИО пользователя")
    doc_ws.write(0, 2, "email")
    doc_ws.write(0, 3, "Логин")
    doc_ws.write(0, 4, "Пароль")

    for user in users_data:
        row = config.EXCEL_START_ROW + count

        doc_ws.write(row, 1, user['FIO'])
        doc_ws.write(row, 2, user['email'])
        doc_ws.write(row, 3, user['login'])
        doc_ws.write(row, 4, user['upass'])
        
        count += 1
    
    doc.close()


def removejirauser(jira, users):
    for user in users:
        try:
            jira.delete_user(user['login'])
            logging.info(f"Пользователь {user['login']} удален.")
            return True
        except JIRAError as err:
            logging.error(err)
            return False

        
def parsepassfile(file):
    users_data = []

    with open(file, 'r', encoding='UTF-8') as File:
        lines = File.readlines()
    File.close()

    for line in lines:
        temp_data = {}
        
        user_fio = re.search(r'\'([а-яёА-ЯЁ_\-\.\(\) ]*)\' ', line)
        user_email = re.search(r' \[([a-zA-Z0-9_\-\.]*@[a-zA-Z0-9_\-]*\.\w+)\] :', line)
        user_login = re.search(r':\s(\w+)\s-', line)
        user_password = re.search(r'- (.*)$', line, re.M)

        if (user_fio and user_email and user_login and user_password):
            temp_data['FIO'] = user_fio.group(1)
            temp_data['login'] = user_login.group(1)
            temp_data['email'] = user_email.group(1)
            temp_data['upass'] = user_password.group(1)
        else:
            print(user_fio, user_email, user_login, user_password)
            logging.error(f"Не удалось пролностью распарсить строку. \n>>>\n{line}<<<")

        users_data.append(temp_data)

    return users_data


if __name__ == "__main__":
    logging.info(f">>>> Старт скрипта {datetime.now()} <<<<")
    print(sys.argv[0])
    users_file = os.path.abspath("users_list.xlsx")
    pass_file = os.path.abspath("password.txt")
    jira = jiraConnect(config.USERS_ADD_BASE_URL)
    

    #  
    # Create users in JIRA from pass file
    #
    # users_data = parsepassfile('password.txt')

    # for user in users_data:
    #     if (addusertojira(jira=jira, user=user)):
    #         logging.info(f"Пользователь {user['FIO']} - {user['login']} создан в JIRA.")
    #     else:
    #         logging.info(f"Произошла ошибка при создании пользователя, либо пользователь уже создан.\n >>> {user['FIO']} - {user['login']} !")

    #
    # Create users in JIRA from Excel file
    #
    # users = readexcelfile(users_file)

    # f = open(pass_file, 'a', encoding='utf-8')
    # count = 0

    # for user in users:
    #     count += 1
    #     try:
    #         if adduser(jira, user, f):
    #             logging.info(f"Для пользователя № {count} : {user['FIO']} создана учетная запись.")
    #     except Exception as e:
    #         logging.error(e)

    # f.close()

    # getjirausers(jira, config.ALL_USERS_CSV_FILE, mode='Activate')
    # addjirausersingroup(jira, config.ACTIVATE_USERS_FILE)
    writexls(pass_file)
    # removejirauser(jira, pass_file)
